"""
Main pipeline with support for both classifiers:
1. HierachicalTaxonomyClassifier (Naive Bayes - original)
2. HierarchicalEnsembleClassifier (Ensemble - new)

Usage:
------
# Train with Naive Bayes (original):
python main_ensemble.py --input data.xlsx --mode train

# Train with Ensemble (new):
python main_ensemble.py --input data.xlsx --mode train --use-ensemble

# Predict with existing model (auto-detects type):
python main_ensemble.py --input data.xlsx --mode predict --model model.joblib
"""

import pandas as pd
import numpy as np
import argparse
from datetime import datetime
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
warnings.filterwarnings('ignore')

from preprocessing import TextPreprocessor, DataPreprocessor
from hierarchical_classifier import HierachicalTaxonomyClassifier
from hierarchical_ensemble_classifier import HierarchicalEnsembleClassifier, EnsembleConfig


class ProcurementTaxonomyPipeline:
    """
    Complete ML pipeline for procurement taxonomy classification.
    
    Now supports TWO classifier types:
    1. Naive Bayes (original) - Fast, simple, good baseline
    2. Ensemble (new) - More accurate, uses multiple models per level
    
    Modes:
    1. TRAIN: Train new model from scratch
    2. UPDATE: Retrain with new data
    3. PREDICT: Predict using existing model
    """
    
    def __init__(self, confidence_threshold=0.5, use_ensemble=False):
        """
        Initialize the pipeline.
        
        Parameters:
        -----------
        confidence_threshold : float
            Minimum confidence for accepting predictions
        use_ensemble : bool
            If True, use HierarchicalEnsembleClassifier
            If False, use HierachicalTaxonomyClassifier (original Naive Bayes)
        """
        self.confidence_threshold = confidence_threshold
        self.use_ensemble = use_ensemble
        self.classifier = None
        self.text_preprocessor = TextPreprocessor()
        self.data_preprocessor = DataPreprocessor(self.text_preprocessor)
        
        print(f"\n📊 Pipeline initialized:")
        print(f"   Classifier type: {'Ensemble (Voting + Stacking + XGBoost)' if use_ensemble else 'Naive Bayes (original)'}")
        print(f"   Confidence threshold: {confidence_threshold}")
    
    def identify_new_rows(self, df,
                          description_col='Συνοπτική Περιγραφή Αντικειμένου Σύμβασης',
                          l1_col='Επίπεδο Κατηγοριοποίησης L.1'):
        """
        Identify which rows need prediction.
        
        Strategy:
        - Rows with L1 labels → LABELED (use for training)
        - Rows without L1 labels → NEW (need prediction)
        """
        print("\n" + "=" * 70)
        print("IDENTIFYING LABELED vs UNLABELED ROWS")
        print("=" * 70)
        
        # Labeled rows
        labeled_mask = df[l1_col].notna()
        labeled_df = df[labeled_mask].copy()
        
        # Unlabeled rows (with descriptions)
        has_description = df[description_col].notna() & (df[description_col].str.strip() != '')
        unlabeled_mask = df[l1_col].isna() & has_description
        unlabeled_df = df[unlabeled_mask].copy()
        
        print(f"\n📊 Data split:")
        print(f"   Total rows: {len(df)}")
        print(f"   Labeled (for training): {len(labeled_df)} ({len(labeled_df)/len(df)*100:.1f}%)")
        print(f"   Unlabeled (for prediction): {len(unlabeled_df)} ({len(unlabeled_df)/len(df)*100:.1f}%)")
        
        if len(labeled_df) < 100:
            print(f"\n⚠️  Warning: Only {len(labeled_df)} labeled rows!")
            print(f"   Model performance may be poor with limited training data.")
        
        return labeled_df, unlabeled_df
    
    def train_model(self, labeled_df, test_size=0.2, random_state=42):
        """
        Train a new model from labeled data.
        
        Steps:
        1. Prepare data
        2. Train classifier (Naive Bayes OR Ensemble)
        3. Evaluate on test set
        
        Returns:
        --------
        dict : Evaluation results
        """
        print("\n" + "=" * 70)
        print(f"TRAINING MODEL ({'ENSEMBLE' if self.use_ensemble else 'NAIVE BAYES'})")
        print("=" * 70)
        
        # Initialize classifier based on type
        if self.use_ensemble:
            print("\n🚀 Using Hierarchical Ensemble Classifier")
            print("   L1: Voting Ensemble (NB + CNB + LR)")
            print("   L2: Stacking Ensemble (NB + CNB + LR + SVC)")
            print("   L3: XGBoost (100 trees, CPU-optimized)")
            
            config = EnsembleConfig()
            self.classifier = HierarchicalEnsembleClassifier(
                config=config,
                text_preprocessor=self.text_preprocessor
            )
        else:
            print("\n📝 Using Naive Bayes Classifier (original)")
            self.classifier = HierachicalTaxonomyClassifier(
                alpha_l1=0.1,
                alpha_l2=0.1,
                alpha_l3=0.1,
                max_features=5000,
                ngram_range=(1, 2),
                min_df=2,
                text_preprocessor=self.text_preprocessor
            )
        
        # Prepare data
        print("\n[1/3] Preparing data...")
        (train_l1, test_l1), (train_l2, test_l2), (train_l3, test_l3) = self.classifier.prepare_data(
            labeled_df,
            test_size=test_size,
            random_state=random_state
        )
        
        # Train
        print("\n[2/3] Training classifier...")
        self.classifier.train(train_l1, train_l2, train_l3)
        
        # Load official taxonomy
        try:
            if hasattr(self, 'current_input_path'):
                self.classifier.load_official_taxonomy(
                    excel_path=self.current_input_path,
                    sheet_name='L1-L2-L3'
                )
        except Exception as e:
            print(f"⚠️  Could not load official taxonomy: {e}")
            print(f"   Continuing with hierarchy learned from training data")
        
        # Evaluate
        print("\n[3/3] Evaluating on test set...")
        results = self.classifier.evaluate(test_l1, test_l2, test_l3)
        
        return results
    
    def predict_new_rows(self, unlabeled_df,
                        description_col='Συνοπτική Περιγραφή Αντικειμένου Σύμβασης'):
        """
        Predict taxonomy for unlabeled rows.
        
        Works with both classifier types (automatically detected).
        """
        print("\n" + "=" * 70)
        print("PREDICTING UNLABELED ROWS")
        print("=" * 70)
        
        if self.classifier is None:
            raise ValueError("No model loaded! Train a model first or load existing one.")
        
        if len(unlabeled_df) == 0:
            print("\n✓ No unlabeled rows to predict!")
            return unlabeled_df
        
        print(f"\n🔮 Predicting {len(unlabeled_df)} rows...")
        
        predictions = []
        
        for i, (idx, row) in enumerate(unlabeled_df.iterrows(), 1):
            # Show progress
            if i % 10 == 0 or i == 1:
                print(f"   Progress: {i}/{len(unlabeled_df)} rows...", end='\r')
            
            description = str(row.get(description_col, ''))
            
            try:
                # Predict (works for both classifier types)
                pred = self.classifier.predict(description, self.confidence_threshold)
                
                if not isinstance(pred, dict):
                    pred = {
                        'l1_pred': None,
                        'l2_pred': None,
                        'l3_pred': None,
                        'l1_conf': 0.0,
                        'l2_conf': 0.0,
                        'l3_conf': 0.0,
                        'combined_conf': 0.0,
                        'accept': False,
                        'reason': 'Prediction returned non-dict'
                    }
            except Exception as e:
                pred = {
                    'l1_pred': None,
                    'l2_pred': None,
                    'l3_pred': None,
                    'l1_conf': 0.0,
                    'l2_conf': 0.0,
                    'l3_conf': 0.0,
                    'combined_conf': 0.0,
                    'accept': False,
                    'reason': f'Error: {str(e)[:100]}'
                }
            
            predictions.append({
                'L1_Prediction': pred['l1_pred'],
                'L2_Prediction': pred['l2_pred'],
                'L3_Prediction': pred['l3_pred'],
                'L1_Confidence': pred['l1_conf'],
                'L2_Confidence': pred['l2_conf'],
                'L3_Confidence': pred['l3_conf'],
                'Combined_Confidence': pred['combined_conf'],
                'Auto_Accept': pred['accept'],
                'Review_Reason': pred['reason']
            })
        
        print(f"\n✓ Completed {len(predictions)} predictions")
        
        # Add predictions to dataframe
        pred_df = pd.DataFrame(predictions, index=unlabeled_df.index)
        result_df = pd.concat([unlabeled_df, pred_df], axis=1)
        
        # Summary
        auto_accept = pred_df['Auto_Accept'].sum()
        manual_review = len(pred_df) - auto_accept
        
        print(f"\n📊 Prediction Summary:")
        print(f"   Total: {len(pred_df)}")
        print(f"   Auto-accept (≥{self.confidence_threshold:.2f}): {auto_accept} ({auto_accept/len(pred_df)*100:.1f}%)")
        print(f"   Manual review (<{self.confidence_threshold:.2f}): {manual_review} ({manual_review/len(pred_df)*100:.1f}%)")
        print(f"   Avg confidence: {pred_df['Combined_Confidence'].mean():.3f}")
        
        return result_df
    
    def merge_results(self, original_df, labeled_df, predicted_df):
        """
        Merge labeled and predicted data back into original dataframe.
        """
        result_df = original_df.copy()
        
        # Initialize prediction columns
        for col in ['L1_Prediction', 'L2_Prediction', 'L3_Prediction',
                   'L1_Confidence', 'L2_Confidence', 'L3_Confidence',
                   'Combined_Confidence', 'Auto_Accept', 'Review_Reason']:
            result_df[col] = np.nan if col != 'Review_Reason' else ''
        
        # Update with predictions
        if len(predicted_df) > 0:
            for col in ['L1_Prediction', 'L2_Prediction', 'L3_Prediction',
                       'L1_Confidence', 'L2_Confidence', 'L3_Confidence',
                       'Combined_Confidence', 'Auto_Accept', 'Review_Reason']:
                if col in predicted_df.columns:
                    result_df.loc[predicted_df.index, col] = predicted_df[col]
        
        print(f"\n📋 Merged results:")
        print(f"   Total rows: {len(result_df)}")
        print(f"   With predictions: {len(predicted_df)}")
        print(f"   Without predictions: {len(labeled_df)}")
        
        return result_df
    
    def save_results_to_excel(self, df, output_path, highlight_low_confidence=True):
        """
        Save results to Excel with highlighting.
        """
        print("\n" + "=" * 70)
        print("SAVING RESULTS TO EXCEL")
        print("=" * 70)
        
        print(f"\n📁 Saving to: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Predictions', index=False)
        
        if highlight_low_confidence:
            self._apply_highlighting(output_path, 'Predictions', self.confidence_threshold)
        
        print(f"\n✓ Results saved successfully!")
    
    def _apply_highlighting(self, excel_path, sheet_name, threshold):
        """
        Apply granular highlighting based on confidence.
        """
        print(f"🎨 Applying granular highlighting...")
        
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]
        
        # Find columns
        column_map = {}
        for idx, cell in enumerate(ws[1], 1):
            col_name = cell.value
            if col_name in ['L1_Prediction', 'L1_Confidence',
                           'L2_Prediction', 'L2_Confidence',
                           'L3_Prediction', 'L3_Confidence']:
                column_map[col_name] = idx
        
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
        highlighted_cells = 0
        highlighted_rows = set()
        
        for row_idx in range(2, ws.max_row + 1):
            for level in ['L1', 'L2', 'L3']:
                conf_col = column_map.get(f'{level}_Confidence')
                pred_col = column_map.get(f'{level}_Prediction')
                
                if conf_col and pred_col:
                    conf_cell = ws.cell(row=row_idx, column=conf_col)
                    
                    if conf_cell.value is not None:
                        try:
                            if float(conf_cell.value) < threshold:
                                ws.cell(row=row_idx, column=pred_col).fill = yellow_fill
                                ws.cell(row=row_idx, column=conf_col).fill = yellow_fill
                                highlighted_cells += 2
                                highlighted_rows.add(row_idx)
                        except (ValueError, TypeError):
                            pass
        
        wb.save(excel_path)
        print(f"   Highlighted {highlighted_cells} cells across {len(highlighted_rows)} rows")
    
    def run(self, input_path, output_path=None, mode='update', model_path=None):
        """
        Run the complete ML pipeline.
        
        Modes:
        - 'train': Train new model
        - 'update': Retrain with new data
        - 'predict': Use existing model
        """
        print("\n" + "=" * 70)
        print("PROCUREMENT TAXONOMY ML PIPELINE")
        print("=" * 70)
        
        self.current_input_path = input_path
        
        # Generate output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            classifier_type = "ensemble" if self.use_ensemble else "nb"
            output_path = f"predictions_{classifier_type}_{timestamp}.xlsx"
        
        print(f"📂 Input: {input_path}")
        print(f"📂 Output: {output_path}")
        
        # Load data
        df = pd.read_excel(input_path, sheet_name='Overall')
        print(f"✓ Loaded {len(df)} rows")
        
        # Identify labeled vs unlabeled
        labeled_df, unlabeled_df = self.identify_new_rows(df)
        
        evaluation_results = None
        
        # TRAIN MODE
        if mode == 'train':
            if len(labeled_df) < 100:
                raise ValueError(f"Insufficient training data! Need ≥100 labeled rows, got {len(labeled_df)}")
            
            evaluation_results = self.train_model(labeled_df)
            
            # Save model
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            classifier_type = "ensemble" if self.use_ensemble else "nb"
            model_save_path = f"taxonomy_classifier_{classifier_type}_{timestamp}.joblib"
            self.classifier.save(model_save_path)
            print(f"\n💾 Model saved to: {model_save_path}")
        
        # UPDATE MODE
        elif mode == 'update':
            if len(labeled_df) < 100:
                print(f"\n⚠️  Warning: Only {len(labeled_df)} labeled rows for training.")
            
            evaluation_results = self.train_model(labeled_df)
            
            # Save updated model
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            classifier_type = "ensemble" if self.use_ensemble else "nb"
            model_save_path = f"taxonomy_classifier_{classifier_type}_{timestamp}.joblib"
            self.classifier.save(model_save_path)
            print(f"\n💾 Updated model saved to: {model_save_path}")
        
        # PREDICT MODE
        elif mode == 'predict':
            if model_path is None:
                raise ValueError("model_path required for 'predict' mode")
            
            print(f"\n📥 Loading model from: {model_path}")
            
            # Auto-detect model type
            model_data = joblib.load(model_path)
            version = model_data.get('version', '1.0')
            
            if 'ensemble' in version:
                print("   Detected: Ensemble model")
                self.classifier = HierarchicalEnsembleClassifier.load(model_path)
                self.use_ensemble = True
            else:
                print("   Detected: Naive Bayes model")
                self.classifier = HierachicalTaxonomyClassifier.load(model_path)
                self.use_ensemble = False
            
            print(f"   ✓ Model loaded successfully!")
        
        else:
            raise ValueError(f"Invalid mode: {mode}")
        
        # Predict on unlabeled rows
        if len(unlabeled_df) > 0:
            predicted_df = self.predict_new_rows(unlabeled_df)
        else:
            predicted_df = pd.DataFrame()
            print("\n✓ No unlabeled rows to predict")
        
        # Merge results
        result_df = self.merge_results(df, labeled_df, predicted_df)
        
        # Save results
        self.save_results_to_excel(result_df, output_path)
        
        print("\n" + "=" * 70)
        print("✅ PIPELINE COMPLETE")
        print("=" * 70)
        
        return result_df, evaluation_results


def main():
    """Main entry point for command-line usage."""
    parser = argparse.ArgumentParser(
        description='Procurement Taxonomy ML Pipeline (Naive Bayes OR Ensemble)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
---------
Train with Naive Bayes (original, fast):
    python main_ensemble.py --input data.xlsx --mode train

Train with Ensemble (new, more accurate):
    python main_ensemble.py --input data.xlsx --mode train --use-ensemble

Predict with existing model (auto-detects type):
    python main_ensemble.py --input data.xlsx --mode predict --model model.joblib

Monthly Update with Ensemble:
    python main_ensemble.py --input new_data.xlsx --mode update --use-ensemble
        """
    )
    
    parser.add_argument('--input', '-i', required=True,
                       help='Path to input Excel file')
    parser.add_argument('--output', '-o', default=None,
                       help='Path to output Excel file (default: auto-generated)')
    parser.add_argument('--mode', '-m', choices=['train', 'update', 'predict'],
                       default='update',
                       help='Mode: train, update, or predict')
    parser.add_argument('--model', default=None,
                       help='Path to existing model (required for predict mode)')
    parser.add_argument('--threshold', '-t', type=float, default=0.5,
                       help='Confidence threshold (default: 0.5)')
    parser.add_argument('--use-ensemble', action='store_true',
                       help='Use Ensemble classifier instead of Naive Bayes')
    
    args = parser.parse_args()
    
    # Validate
    if args.mode == 'predict' and args.model is None:
        parser.error("--model is required when mode is 'predict'")
    
    # Run pipeline
    pipeline = ProcurementTaxonomyPipeline(
        confidence_threshold=args.threshold,
        use_ensemble=args.use_ensemble
    )
    
    result_df, evaluation_results = pipeline.run(
        input_path=args.input,
        output_path=args.output,
        mode=args.mode,
        model_path=args.model
    )


if __name__ == '__main__':
    main()
