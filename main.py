import pandas as pd
import numpy as np
import argparse
from datetime import datetime
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
warnings.filterwarnings('ignore')

from preprocessing import TextPreprocessor, DataPreprocessor
from hierarchical_classifier import HierachicalTaxonomyClassifier

class ProcurementTaxonomyPipeline:
    """
    Complete ML pipeline for procurement taxonomy classification.

    Handles three modes:
    1. TRAIN: Train new model from scratch
    2. UPDATE: Retrain with new data (monthly updates)
    3. PREDICT: Predict using existing model
    """

    def __init__(self, confidence_threshold=0.5):
        """
        Initialize the pipeline.

        Parameters:
        -----------
        confidence_threshold: float
            Minimum confidence for accepting predictions (default: 0.5)
            Below this, predictions are flagged for manual review
        """

        self.confidence_threshold = confidence_threshold
        self.classifier = None
        self.text_preprocessor = TextPreprocessor()
        self.data_preprocessor = DataPreprocessor(self.text_preprocessor)


    def identify_new_rows(self, df,
                          description_col='Συνοπτική Περιγραφή Αντικειμένου Σύμβασης', 
                          l1_col='Επίπεδο Κατηγοριοποίησης L.1'):
        
        """
        Identify which rows need prediction.

        Strategy:
        --------
        1. Rows with missing L1 labels -> NEW (need prediction)
        2. Rows with existing L1 labels -> LABELED (use for training)

        Parameters:
        ----------
        df : pd.DataFrame
            Input dataframe

        description_col : str
            Column with text descriptions

        l1_col : str
            Column with L1 labels

        Returns:
        -------
        tuple : (labeled_df, unlabeled_df)
            - labeled_df: Rows with existing labels (for training)
            - unlabeled_df: Rows without labels (need prediction)
        """

        print("\n" + "="*70)
        print("IDENTIFYING NEW ROWS")
        print("="*70)

        # Rows with L1 labels = already processed (use for training)
        labeled_mask = df[l1_col].notna()
        labeled_df = df[labeled_mask].copy()

        # Rows without L1 labels = new (need prediction)
        unlabeled_mask = df[l1_col].isna()
        unlabeled_df = df[unlabeled_mask].copy()

        # Also check for rows with descriptions but no labels
        has_description = df[description_col].notna() & (df[description_col].str.strip() != '')
        unlabeled_df = df[unlabeled_mask & has_description].copy()

        print(f"\n Data Split:")
        print(f"    Total rows: {len(df)}")
        print(f"    Labeled rows (for training): {len(labeled_df)} ({len(labeled_df)}/{len(df)*100:.1f}%)")
        print(f"    Unlabeled rows (for prediction): {len(unlabeled_df)} ({len(unlabeled_df)}/{len(df)*100:.1f}%)")

        if len(unlabeled_df) == 0:
            print("\n No unlabeled rows found!")

        if len(labeled_df) < 100:
            print(f"\n Warning: very few labeled rows ({len(labeled_df)})!")
            print(f"    Model performance may be poor with limited training data.")

        return labeled_df, unlabeled_df
    

    def train_model(self, labeled_df, test_size=0.2, random_state=42):
        """
        Train a new model from labeled data.

        Steps:
        ------
        1. Prepare data (preprocess, split train/test)
        2. Train hierarchical classifier
        3. Evaluate on test set
        4. Save model

        Parameters:
        -----------
        labeled_df : pd.DataFrame
            Data with existing labels
        test_size : float
            Proportion for test set
        random_state : int
            Random seed

        Returns:
        --------
        dict : Evaluation results
        """

        print("\n" + "="*70)
        print("TRAINING MODEL")
        print("="*70)

        # Initialize classifier
        self.classifier = HierachicalTaxonomyClassifier(
            alpha_l1=0.1,
            alpha_l2=0.1,
            alpha_l3=0.1,
            max_features=5000,
            ngram_range=(1,2),
            min_df=2,
            text_preprocessor=self.text_preprocessor
        )

        # Prepare data
        print("\n[1/3] Preparing data...")
        (train_l1, test_l1), (train_l2, test_l2), (train_l3, test_l3) = self.classifier.prepare_data(
            labeled_df, 
            test_size=test_size, 
            random_state=random_state)

        # Train
        print("\n[2/3] Training classifier...")
        self.classifier.train(train_l1, train_l2, train_l3)

        # Load official taxonomy (overrides learned hierarchy)
        try:
            if hasattr(self, 'current_input_path'):
                self.classifier.load_official_taxonomy(
                    excel_path=self.current_input_path,
                    sheet_name='L1-L2-L3'
                )
        except Exception as e:
            print(f"    Could not load official taxonomy: {e}")
            print(f"    Continuing with hierarchy learned from training data")

        # Evaluate
        print("\n[3/3] Evaluating on test set...")
        results = self.classifier.evaluate(test_l1, test_l2, test_l3)

        return results

    def predict_new_rows(self, unlabeled_df, 
                        description_col='Συνοπτική Περιγραφή Αντικειμένου Σύμβασης'):
        """
        Predict taxonomy for new/unlabeled rows.

        Parameteres:
        ------------
        unlabeled_df : pd.DataFrame
            Rows that need predictions
        description_col : str
            Column with text description

        Returns:
        --------
        pd.DataFrame : Original dataframe with predictions added
        """

        print("\n" + "="*70)
        print("PREDICTING NEW ROWS")
        print("="*70)

        if self.classifier is None:
            raise ValueError("No model loaded! Train a model first or load existing one.")
        
        if len(unlabeled_df) == 0:
            print("\n No unlabeled rows to predict!")
            return unlabeled_df
        
        print(f"\n Predicting {len(unlabeled_df)} rows...")

        # Initialize result columns
        predictions = []
        error_count = 0
        success_count = 0

        for i, (idx, row) in enumerate(unlabeled_df.iterrows(), 1):
            # Show progress every 10 rows
            if i%10 == 0 or i == 1:
                print(f"    Progress: {i}/{len(unlabeled_df)} rows...", end='\r')

            description = row[description_col]

            if pd.isna(description):
                description = ""
            else:
                description = str(description)

            # Debug: Show first prediction attempt
            if i == 1:
                print(f"\n  [DEBUG] First row:")
                print(f"    Idx: {idx}")
                print(f"    Description: '{description[:100]}...'")
                print(f"    Classifier exists: {self.classifier is not None}")

            try:
                # Predict
                pred = self.classifier.predict(description, self.confidence_threshold)

                if i == 1:
                    print(f"    Prediction type: {type(pred)}")
                    print(f"    Prediction keys: {pred.keys() if isinstance(pred, dict) else 'N/A'}")
                    if isinstance(pred, dict):
                        print(f"    L1: {pred.get('l1_pred')}")
                        print(f"    L2: {pred.get('l2_pred')}")
                        print(f"    L3: {pred.get('l3_pred')}")
                        print(f"    Confidence: {pred.get('combined_conf')}")

                if not isinstance(pred, dict):
                    pred = {
                        'l1_pred': None,
                        'l2_pred': None,
                        'l3_pred': None,
                        'l1_conf': 0.0,
                        'l2_conf': 0.0,
                        'l3_conf': 0.0,
                        'combined_conf': 0.0,
                        'accept': False,
                        'reason': 'Prediction return non-dict'
                    }
            except Exception as e:
                error_count += 1
                if error_count <= 5:
                    print(f"\nRow {idx}: {str(e)[:100]}")

                # Fallback on any error
                pred = {
                        'l1_pred': None,
                        'l2_pred': None,
                        'l3_pred': None,
                        'l1_conf': 0.0,
                        'l2_conf': 0.0,
                        'l3_conf': 0.0,
                        'combined_conf': 0.0,
                        'accept': False,
                        'reason': f'Error: {str(e)}'
                    }

            predictions.append({
                'L1_Prediction': pred['l1_pred'],
                'L2_Prediction': pred['l2_pred'],
                'L3_Prediction': pred['l3_pred'],
                'L1_Confidence': pred['l1_conf'],
                'L2_Confidence': pred['l2_conf'],
                'L3_Confidence': pred['l3_conf'],
                'Combined_Confidence': pred['combined_conf'],
                'Auto_Accept': pred['accept'],
                'Review_Reason': pred['reason']
            })

        print(f"\nCompleted {len(predictions)} predictions")
        print(f"    Successful: {success_count}")
        if error_count > 0:
            print(f"{error_count} predictions failed (see errors above)")

        # Debug: Show sample predictions
        if len(predictions) > 0:
            print(f"\n  [DEBUG] First 3 predictions:")
            for i, p in enumerate(predictions[:3], 1):
                print(f"    {i}. L1={p['L1_Prediction']}, L2={p['L2_Prediction']}, L3={p['L3_Prediction']}")
                print(f"        Conf={p['Combined_Confidence']}, Reason='{p['Review_Reason']}'")

        # Add prediction to dataframe
        pred_df = pd.DataFrame(predictions, index=unlabeled_df.index)
        result_df = pd.concat([unlabeled_df, pred_df], axis=1)

        # Summary statistics
        auto_accept_count = pred_df['Auto_Accept'].sum()
        manual_review_count = len(pred_df) - auto_accept_count

        print(f"\n Prediction Summary")
        print(f"    Total predictions: {len(pred_df)}")
        print(f"    Auto-accept(>={self.confidence_threshold:.2f}): {auto_accept_count} ({auto_accept_count/len(pred_df)*100:.1f}%)")
        print(f"    Manual review(<{self.confidence_threshold:.2f}): {auto_accept_count} ({manual_review_count/len(pred_df)*100:.1f}%)")
        print(f"    Average confidence: {pred_df['Combined_Confidence'].mean():.3f}")
        print(f"    Min confidence: {pred_df['Combined_Confidence'].min():.3f}")
        print(f"    Max confidence: {pred_df['Combined_Confidence'].max():.3f}")

        return result_df
        
    def merge_results(self, original_df, labeled_df, predicted_df):
        """
        Merge labeled and predicted data back into original dataframe.

        Strategy:
        ---------
        1. Keep all original columns
        2. For labeled rows: Keep existing L1/L2/L3
        3. For predicted rows: Add L1/L2/L3 predictions + confidence scores

        Parameters:
        -----------

        Returns:
        --------
        pd.DataFrame : Complete merged dataframe
        """

        result_df = original_df.copy()

        result_df['L1_Prediction'] = np.nan
        result_df['L2_Prediction'] = np.nan
        result_df['L3_Prediction'] = np.nan
        result_df['L1_Confidence'] = np.nan
        result_df['L2_Confidence'] = np.nan
        result_df['L3_Confidence'] = np.nan
        result_df['Combined_Confidence'] = np.nan
        result_df['Auto_Accept'] = np.nan
        result_df['Review_Reason'] = ''

        # Update with predictions
        if len(predicted_df) > 0:
            print(f"\n[DEBUG] Merging predictions:")
            print(f"    predicted_df shape: {predicted_df.shape}")
            print(f"    predicted_df columns: {list(predicted_df.columns)}")
            print(f"    predicted_df index: {predicted_df.index[:5].tolist()}")

            # Check if prediction columns have values
            for col in ['L1_Prediction', 'L2_Prediction', 'L3_Prediction']:
                if col in predicted_df.columns:
                    non_null = predicted_df[col].notna().sum()
                    print(f"    {col}: {non_null}/{len(predicted_df)} non-null values")
                    if non_null > 0:
                        print(f"    Sample: {predicted_df[col].dropna().head(2).tolist()}")

            # Update with predictions
            if len(predicted_df) > 0:
                for col in ['L1_Prediction', 'L2_Prediction', 'L3_Prediction',
                            'L1_Confidence', 'L2_Confidence', 'L3_Confidence',
                            'Combined_Confidence', 'Auto_Accept', 'Review_Reason']:
                    if col in predicted_df.columns:
                        result_df.loc[predicted_df.index, col] = predicted_df[col]

        else:
            print(f"\n[DEBUG] No predictions to merge (predicted_df is empty)")

        print(f"\nMerged results:")
        print(f"    Total rows: {len(result_df)}")
        print(f"    Rows with predictions: {len(predicted_df)}")
        print(f"    Rows without predictions: {len(labeled_df)}")

        return result_df

    def save_results_to_excel(self, df, output_path, evaluation_results=None, highlight_low_confidence=True):
        """
        Save results to Excel with conditional formatting.

        Features:
        ---------
        1. Save all data to Excel
        2. Highlight rows with low confidence in yellow
        3. Add evaluation metrics sheet with business explanations

        """

        print("\n" + "="*70)
        print("SAVING RESUTLS TO EXCEL")
        print("="*70)

        print(f"\nSaving to: {output_path}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Predictions', index=False)

            # Evaluation metrics sheet (if available)
            if evaluation_results is not None:
                metrics_df = self._create_evaluation_metrics_sheet(df, evaluation_results)
                metrics_df.to_excel(writer, sheet_name='Evaluation Metrics', index=False)

        if highlight_low_confidence:
            self._apply_highlighting(output_path, 'Predictions', self.confidence_threshold)

        sheets_list = ['Predictions']
        if evaluation_results is not None:
            sheets_list.append('Evaluation Metrics')

        print(f"\nResults saved successfully!")
        print(f"    File: {output_path}")
        print(f"    Sheets: {', '.join(sheets_list)}")

    def _create_evaluation_metrics_sheet(self, df, evaluation_results):
        """
        Create evaluation metrics sheet with business explanations.

        Parameters:
        -----------
        df : pd.DataFrame
            Predictions dataframe
        evaluation_results : dict
            Model evaluation results from training

        Returns:
        --------
        pd.DataFrame : Formatted metrics data
        """

        # Calculate prediction statistics
        has_predictions = df['L1_Prediction'].notna()
        total_predictions = has_predictions.sum()

        if total_predictions > 0:
            auto_accept_count = df.loc[has_predictions, 'Auto_Accept'].sum()
            auto_accept_pct = (auto_accept_count / total_predictions) * 100
            manual_review_count = total_predictions - auto_accept_count
            manual_review_pct = (manual_review_count / total_predictions) * 100
            avg_confidence = df.loc[has_predictions, 'Combined_Confidence'].mean()
            min_confidence = df.loc[has_predictions, 'Combined_Confidence'].min()
            max_confidence = df.loc[has_predictions, 'Combined_Confidence'].max()
        else:
            auto_accept_count = 0
            auto_accept_pct = 0
            manual_review_count = 0
            manual_review_pct = 0
            avg_confidence = 0
            min_confidence = 0
            max_confidence = 0

        # Build metrics data
        rows = []

        # Section 1: Prediction Summary
        rows.append({'Metric': '=== PREDICTION SUMMARY ===', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({
            'Metric': 'Total Predictions',
            'Value': int(total_predictions),
            'Explanation': 'Number of rows classified by the model'
        })
        rows.append({
            'Metric': 'Auto-Accept Count',
            'Value': int(auto_accept_count),
            'Explanation': f'Predictions with confidence >= {self.confidence_threshold:.2f} (ready to use)'
        })
        rows.append({
            'Metric': 'Auto-Accept Percentage',
            'Value': f'{auto_accept_pct:.1f}%',
            'Explanation': 'Proportion of predictions that meet confidence threshold'
        })
        rows.append({
            'Metric': 'Manual Review Count',
            'Value': int(manual_review_count),
            'Explanation': f'Predictions with confidence < {self.confidence_threshold:.2f} (need human review)'
        })
        rows.append({
            'Metric': 'Manual Review Percentage',
            'Value': f'{manual_review_pct:.1f}%',
            'Explanation': 'Proportion of predictions flagged for manual verification'
        })
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({
            'Metric': 'Average Confidence',
            'Value': f'{avg_confidence:.3f}',
            'Explanation': 'Mean confidence across all predictions (higher is better)'
        })
        rows.append({
            'Metric': 'Min Confidence',
            'Value': f'{min_confidence:.3f}',
            'Explanation': 'Lowest confidence score (identifies most uncertain predictions)'
        })
        rows.append({
            'Metric': 'Max Confidence',
            'Value': f'{max_confidence:.3f}',
            'Explanation': 'Highest confidence score (best model certainty)'
        })

        # Section 2: Model Performance
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '=== MODEL PERFORMANCE (Test Set) ===', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})

        # L1 metrics
        rows.append({
            'Metric': 'L1 Accuracy',
            'Value': f"{evaluation_results['l1']['accuracy']:.1%}",
            'Explanation': 'Top-level category classification accuracy (how often L1 predictions are correct)'
        })
        rows.append({
            'Metric': 'L1 F1-Score (Weighted)',
            'Value': f"{evaluation_results['l1']['f1_weighted']:.3f}",
            'Explanation': 'L1 quality score balancing precision and recall across all categories (0-1, higher is better)'
        })
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})

        # L2 metrics
        rows.append({
            'Metric': 'L2 Accuracy',
            'Value': f"{evaluation_results['l2']['accuracy']:.1%}",
            'Explanation': 'Mid-level category classification accuracy (L2 is harder than L1 due to more classes)'
        })
        rows.append({
            'Metric': 'L2 F1-Score (Weighted)',
            'Value': f"{evaluation_results['l2']['f1_weighted']:.3f}",
            'Explanation': 'L2 quality score balancing precision and recall (lower than L1 is expected)'
        })
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})

        # L3 metrics
        rows.append({
            'Metric': 'L3 Accuracy',
            'Value': f"{evaluation_results['l3']['accuracy']:.1%}",
            'Explanation': 'Fine-grained category classification accuracy (L3 is most challenging level)'
        })
        rows.append({
            'Metric': 'L3 F1-Score (Weighted)',
            'Value': f"{evaluation_results['l3']['f1_weighted']:.3f}",
            'Explanation': 'L3 quality score balancing precision and recall (lowest scores are normal)'
        })

        # Section 3: Metric Definitions
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '=== METRIC DEFINITIONS ===', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({
            'Metric': 'Confidence Score',
            'Value': '0.0 - 1.0',
            'Explanation': 'Model certainty in prediction. Combines L1, L2, L3 probabilities. Higher = more certain.'
        })
        rows.append({
            'Metric': 'Accuracy',
            'Value': '% correct',
            'Explanation': 'Percentage of test predictions that exactly match true labels. Simple correctness metric.'
        })
        rows.append({
            'Metric': 'F1-Score',
            'Value': '0.0 - 1.0',
            'Explanation': 'Harmonic mean of precision (% of predictions that are correct) and recall (% of true labels found). Better for imbalanced data.'
        })
        rows.append({
            'Metric': 'Weighted F1',
            'Value': 'avg by class',
            'Explanation': 'F1-score averaged across all categories, weighted by number of samples in each category.'
        })
        rows.append({
            'Metric': 'Highlighting',
            'Value': 'Yellow cells',
            'Explanation': f'Only taxonomy levels (L1/L2/L3) with confidence < {self.confidence_threshold:.2f} are highlighted. Allows granular review.'
        })

        # Section 4: Recommendations
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '=== RECOMMENDATIONS ===', 'Value': '', 'Explanation': ''})
        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})

        if auto_accept_pct >= 70:
            recommendation = 'GOOD: High auto-accept rate indicates model is performing well. Most predictions can be used directly.'
        elif auto_accept_pct >= 50:
            recommendation = 'MODERATE: Acceptable performance but significant manual review needed. Consider lowering threshold or improving training data.'
        else:
            recommendation = 'LOW: High manual review rate suggests model uncertainty. Recommended actions: (1) Add more training data, (2) Lower confidence threshold, (3) Review data quality.'

        rows.append({
            'Metric': 'Overall Assessment',
            'Value': f'{auto_accept_pct:.0f}% auto-accept',
            'Explanation': recommendation
        })

        rows.append({'Metric': '', 'Value': '', 'Explanation': ''})
        rows.append({
            'Metric': 'Next Steps',
            'Value': 'Review highlighted',
            'Explanation': 'Focus on yellow-highlighted cells in Predictions sheet. These need human verification before use.'
        })

        return pd.DataFrame(rows)

    def _apply_highlighting(self, excel_path, sheet_name, threshold):
        """
        Apply granular highlighting to prediction and confidence columns based on individual level confidence.

        Highlights only the specific taxonomy level (L1/L2/L3) predictions and their confidence scores
        that fall below the threshold, rather than highlighting entire rows.
        """

        print(f"Applying granular highlighting...")

        # Load workbook
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        # Find all relevant columns
        column_map = {}
        for idx, cell in enumerate(ws[1], 1):
            col_name = cell.value
            if col_name in ['L1_Prediction', 'L1_Confidence',
                           'L2_Prediction', 'L2_Confidence',
                           'L3_Prediction', 'L3_Confidence']:
                column_map[col_name] = idx

        if len(column_map) == 0:
            print("    Could not find prediction/confidence columns for highlighting")
            return

        # Yellow fill for low confidence
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

        highlighted_cells = 0
        highlighted_rows = set()

        for row_idx in range(2, ws.max_row + 1):  # Skip header
            row_highlighted = False

            # Check each level (L1, L2, L3)
            for level in ['L1', 'L2', 'L3']:
                conf_col = column_map.get(f'{level}_Confidence')
                pred_col = column_map.get(f'{level}_Prediction')

                if conf_col and pred_col:
                    confidence_cell = ws.cell(row=row_idx, column=conf_col)

                    if confidence_cell.value is not None:
                        try:
                            confidence = float(confidence_cell.value)
                            if confidence < threshold:
                                # Highlight both the prediction and confidence cells for this level
                                ws.cell(row=row_idx, column=pred_col).fill = yellow_fill
                                ws.cell(row=row_idx, column=conf_col).fill = yellow_fill
                                highlighted_cells += 2
                                row_highlighted = True
                        except (ValueError, TypeError):
                            pass

            if row_highlighted:
                highlighted_rows.add(row_idx)

        # Save workbook
        wb.save(excel_path)
        print(f"    Highlighted {highlighted_cells} cells across {len(highlighted_rows)} rows")
        print(f"    (Only taxonomy levels with confidence < {threshold:.2f} are highlighted)")

    def run(self, input_path, output_path=None, mode='update', model_path=None):
        """
        Run the complete ML pipeline.

        Parameters:
        -----------

        Returns:
        --------
        tuple : (result_df, evaluation_results)
        """

        print("\n" + "="*70)
        print("PROCUREMENT TAXONOMY ML PIPELINE")
        print("="*70)

        # Store input path for use in train_model
        self.current_input_path = input_path

        # Generate output path if not provided
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"predictions_{timestamp}.xlsx"
        print(f"Output: {output_path}")

        # Load data
        df = pd.read_excel(input_path, sheet_name='Overall')
        print(f"    Loaded {len(df)} rows")

        # Identify labeled vs unlabeled rows
        labeled_df, unlabeled_df = self.identify_new_rows(df)

        evaluation_results = None

        # TRAIN MODE: Train new model from scratch
        if mode == 'train':
            if len(labeled_df) < 100:
                raise ValueError(f"Insufficient training data! Need at least 100 labeled rows, got {len(labeled_df)}")
            
            evaluation_results = self.train_model(labeled_df)

            # Save model
            model_save_path = f"taxonomy_classifier_{datetime.now().strftime('%Y%m%d_%H%M%S')}.joblib"
            self.classifier.save(model_save_path)
            print(f"Model saved to: {model_save_path}")

        # UPDATE MODE: Retrain with new + old data
        elif mode == 'update':
            if len(labeled_df) < 100:
                print(f"\n Warning: Only {len(labeled_df)} labeled rows available for training.")
                print(f"     Model performance may be poor.")

            evaluation_results = self.train_model(labeled_df)

            # Save updated model
            model_save_path = f"taxonomy_classifier_{datetime.now().strftime('%Y%m%d_%H%M%S')}.joblib"
            self.classifier.save(model_save_path)
            print(f"Updated model saved to: {model_save_path}")
        
        # PREDICT MODE: Use existing model
        elif mode == 'predict':
            if model_path is None:
                raise ValueError("model_path required for 'predict' mode")
            
            print(f"Loading model from: {model_path}")
            self.classifier = HierachicalTaxonomyClassifier.load(model_path)
            print(f"    Model loaded successfully!")

        else:
            raise ValueError(f"Invalid mode: {mode}. Must be 'train', 'update', or 'predict'")
        
        # Predict on unlabeled rows
        if len(unlabeled_df) > 0:
            predicted_df = self.predict_new_rows(unlabeled_df)
        else:
            predicted_df = pd.DataFrame()
            print("\nNo unlabeled rows to predict. All rows already have labels.")

        # Merge results
        result_df = self.merge_results(df, labeled_df, predicted_df)

        # Save results
        self.save_results_to_excel(result_df, output_path, evaluation_results)

        print("\n" + "="*70)
        print("PIPELINE COMPLETE")
        print("="*70)

        return result_df, evaluation_results

def main():
    """Main entry point for command-line usage."""
    parser = argparse.ArgumentParser(
        description='Procurement Taxonomy ML Pipeline',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
                Examples:
                ---------
                Monthly Update(Recommended):
                    python main.py --input new_data.xlsx --mode update

                First Time Training:
                    python main.py --input data.xlsx --mode train

                Prediction Only:
                    python main.py --input data.xlsx --mode predict --model model.joblib

                Custom Confidence Threshold:
                    python main.py --input data.xlsx --mode update --threshold 0.6
                """
    )

    parser.add_argument('--input', '-i', required=True,
                        help='Path to input Excel file')
    parser.add_argument('--output', '-o', default=None,
                        help='Path to output Excel file (default: auto-generated)')
    parser.add_argument('--mode', '-m', choices=['train', 'update', 'predict'], 
                        default='update',
                        help='Mode: train (new model), update (retrain), predict (existing model)')
    parser.add_argument('--model', default=None,
                        help='Path to existing model (required for predict mode)')
    parser.add_argument('--threshold', '-t', type=float, default=0.5,
                        help='Confidence threshold for auto-accept (default: 0.5)')
    
    args = parser.parse_args()

    # Validate arguments
    if args.mode == 'predict' and args.model is None:
        parser.error("--model is required when mode is 'predict'")

    # Run pipeline
    pipeline = ProcurementTaxonomyPipeline(confidence_threshold=args.threshold)
    result_df, evaluation_results = pipeline.run(
        input_path=args.input,
        output_path=args.output,
        mode=args.mode,
        model_path=args.model
    )


if __name__ == '__main__':
        main()
